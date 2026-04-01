from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_PATH = ROOT / "outputs" / "reports" / "fault_sheet_replica.xlsx"


def apply_border(ws, cell_range: str, *, medium_outline: bool = False) -> None:
    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")
    rows = ws[cell_range]
    row_count = len(rows)
    col_count = len(rows[0])
    for r_idx, row in enumerate(rows):
        for c_idx, cell in enumerate(row):
            left = medium if c_idx == 0 and medium_outline else thin
            right = medium if c_idx == col_count - 1 and medium_outline else thin
            top = medium if r_idx == 0 and medium_outline else thin
            bottom = medium if r_idx == row_count - 1 and medium_outline else thin
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)


def style_cell(cell, *, bold: bool = False, size: int = 11, align: str = "left") -> None:
    cell.font = Font(name="Times New Roman", size=size, bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Fault Sheet"
    ws.sheet_view.showGridLines = False

    widths = {
        "A": 11,
        "B": 22,
        "C": 18,
        "D": 18,
        "E": 20,
        "F": 14,
        "G": 18,
        "H": 14,
        "I": 12,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    for row in range(1, 25):
        ws.row_dimensions[row].height = 22
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[4].height = 24
    ws.row_dimensions[5].height = 26
    ws.row_dimensions[16].height = 28
    ws.row_dimensions[21].height = 28

    green_fill = PatternFill(fill_type="solid", fgColor="1B8A5A")

    for cell_ref in ["A1", "B1", "C1", "D1", "E1", "F1", "G1", "H1"]:
        ws[cell_ref].fill = green_fill

    ws.merge_cells("B2:E2")
    ws["B2"] = "BOTSWANA POWER CORPORATION"
    style_cell(ws["B2"], bold=True, align="center")

    ws["A3"] = "FAULT NO:"
    style_cell(ws["A3"], bold=True)
    ws.merge_cells("B3:C3")
    ws["B3"] = "060/22"
    style_cell(ws["B3"], bold=True, align="center")
    apply_border(ws, "A3:C3", medium_outline=True)

    ws["A5"] = "DAY/DATE"
    ws["B5"] = "Monday 25 February 2024"
    ws["D5"] = "WEATHER"
    ws["E5"] = "CLEAR"
    ws["F5"] = "VOLTAGE LEVEL"
    ws["G5"] = "132/11k"
    ws["H5"] = "AREA:"
    ws["I5"] = "South"

    for ref in ["A5", "D5", "F5", "H5"]:
        style_cell(ws[ref], bold=True)
    for ref in ["B5", "E5", "G5", "I5"]:
        style_cell(ws[ref], bold=True)

    ws["A6"] = "TIME"
    ws["B6"] = "APPARATUS TRIPPED"
    ws.merge_cells("D6:E6")
    ws["D6"] = "PROTECTION INDICATIONS"
    ws["F6"] = "TIME RECLOSED"
    ws["G6"] = "REPORTED BY"
    ws["H6"] = "BREAKER\nSEQUENCE"

    for ref in ["A6", "B6", "D6", "F6", "G6", "H6"]:
        style_cell(ws[ref], bold=True)

    ws["B8"] = "Gab East 132/11 kV T1B : HV CB 110B"
    ws["D8"] = "did not trip"
    ws["B9"] = ": LV CB 1H0B"
    ws["D9"] = "did not trip"
    ws["A10"] = "12h52"
    ws["B10"] = "Gab East 132/11 kV T1A : HV CB 110A"
    ws["D10"] = "O/C & E/F : A & B To Ground"
    ws["F10"] = "14h08"
    ws["G10"] = "SCADA/R. Galetshets"
    ws["H10"] = "Tripped"
    ws["B11"] = ": LV CB 1H0A"
    ws["D11"] = "O/C & E/F : A & B To Ground"
    ws["F11"] = "14h14"
    ws["G11"] = "SCADA/R. Galetshets"
    ws["H11"] = "Tripped"

    for row in range(8, 16):
        for col in range(1, 9):
            style_cell(ws.cell(row=row, column=col), size=10)

    ws.merge_cells("A16:I16")
    ws["A16"] = "OPERATIONS CARRIED OUT"
    style_cell(ws["A16"], bold=True)

    ws.merge_cells("A18:I18")
    ws["A18"] = (
        "CSS requested to close one feeder at a time and 132/11 kV Transformer 1A closed first "
        "and Transformer 2A followed and they both held."
    )
    style_cell(ws["A18"], size=10)

    ws.merge_cells("A21:I21")
    ws["A21"] = "LINE INSPECTION ETC"
    style_cell(ws["A21"], bold=True)

    ws.merge_cells("A22:I22")
    ws["A22"] = (
        "CSS advised fault was due to one of the feeders CB failed to trip during fault, "
        "hence through fault."
    )
    style_cell(ws["A22"], size=10)

    apply_border(ws, "A5:I15", medium_outline=True)
    apply_border(ws, "A16:I18", medium_outline=True)
    apply_border(ws, "A21:I22", medium_outline=True)

    wb.save(OUTPUT_PATH)
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
